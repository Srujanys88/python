from openpyxl import Workbook
from openpyxl import *


wb = Workbook()
sheet = wb.active
sheet.title = "language"
wb.create_sheet(title="capital")
lang = ["kannada","telugu","tamilu"]
state = ["Karnataka","Telangana", "Tamilnadu"]
capital = ["Bengaluru", "Hydrabad","chennai"]
code = ["KA","UTS", "TN"]
sheet.cell(row = 1, column =1).value = "state"
sheet.cell(row = 1, column =2).value = "language"
sheet.cell(row = 1, column =3).value = "code"

for i in range(2,5):
  sheet.cell(row = i, column =1).value = state[i-2]
  sheet.cell(row = i, column =2).value = lang[i-2]
  sheet.cell(row = i, column =3).value = code[i-2]

wb.save("cs.xlsx")
sheet = wb["capital"]
sheet.cell(row = 1, column =1).value = "state"
sheet.cell(row = 1, column =2).value = "language"
sheet.cell(row = 1, column =3).value = "code"

for i in range(2,5):
  sheet.cell(row = i, column =1).value = state[i-2]
  sheet.cell(row = i, column =2).value = capital[i-2]
  sheet.cell(row = i, column =3).value = code[i-2]

wb.save("cs.xlsv")

srchcode = input("Enter state code for finding capital")
for i in range(2,5):
  data = sheet.cell(row = i, column = 3).value
  if data == srchcode:
    print ("capital of ",srchcode, "is", sheet.cell(row=i,column=2).value )

sheet = wb['language']
srchcode = input("Enter the state code for finding language")

for i in range(2,5):
  data = sheet.cell(row = i, column = 3).value
  if data == srchcode:
    print ("capital of ",srchcode, "is", sheet.cell(row=i,column=2).value )
wb.close()